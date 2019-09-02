#Joshua Van Daalen

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')

for x in wb.sheetnames:
    ws = wb[x]
    for col in ws.iter_cols():
        col[0].value = col[0].value.strip() #Trim whitespace from header cells
    ws.delete_rows(2,1) #Remove Second column with no data
